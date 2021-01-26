import bs4
import requests
import urllib.request
import random
import time
import json
import openpyxl


def requests_headers():  # 构造请求头池
    head_connection = ['Keep-Alive', 'close']
    head_accept = ['text/html,application/xhtml+xml,*/*']
    head_accept_language = ['zh-CN,fr-FR;q=0.5', 'en-US,en;q=0.8,zh-Hans-CN;q=0.5,zh-Hans;q=0.3']
    head_user_agent = ['Mozilla/5.0 (Windows NT 6.3; WOW64; Trident/7.0; rv:11.0) like Gecko',
                       'Mozilla/5.0 (Windows NT 5.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/28.0.1500.95 Safari/537.36',
                       'Mozilla/5.0 (Windows NT 6.1; WOW64; Trident/7.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; Media Center PC 6.0; .NET4.0C; rv:11.0) like Gecko)',
                       'Mozilla/5.0 (Windows; U; Windows NT 5.2) Gecko/2008070208 Firefox/3.0.1',
                       'Mozilla/5.0 (Windows; U; Windows NT 5.1) Gecko/20070309 Firefox/2.0.0.3',
                       'Mozilla/5.0 (Windows; U; Windows NT 5.1) Gecko/20070803 Firefox/1.5.0.12',
                       'Opera/9.27 (Windows NT 5.2; U; zh-cn)',
                       'Mozilla/5.0 (Macintosh; PPC Mac OS X; U; en) Opera 8.0',
                       'Opera/8.0 (Macintosh; PPC Mac OS X; U; en)',
                       'Mozilla/5.0 (Windows; U; Windows NT 5.1; en-US; rv:1.8.1.12) Gecko/20080219 Firefox/2.0.0.12 Navigator/9.0.0.6',
                       'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.1; Win64; x64; Trident/4.0)',
                       'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.1; Trident/4.0)',
                       'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.1; WOW64; Trident/6.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; Media Center PC 6.0; InfoPath.2; .NET4.0C; .NET4.0E)',
                       'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.1 (KHTML, like Gecko) Maxthon/4.0.6.2000 Chrome/26.0.1410.43 Safari/537.1 ',
                       'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.1; WOW64; Trident/6.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; Media Center PC 6.0; InfoPath.2; .NET4.0C; .NET4.0E; QQBrowser/7.3.9825.400)',
                       'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:21.0) Gecko/20100101 Firefox/21.0 ',
                       'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.1 (KHTML, like Gecko) Chrome/21.0.1180.92 Safari/537.1 LBBROWSER',
                       'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.1; WOW64; Trident/6.0; BIDUBrowser 2.x)',
                       'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/536.11 (KHTML, like Gecko) Chrome/20.0.1132.11 TaoBrowser/3.0 Safari/536.11']
    header = {
        'Connection': head_connection[random.randrange(0, len(head_connection))],
        'Accept': head_accept[0],
        'Accept-Language': head_accept_language[random.randrange(0, len(head_accept_language))],
        'User-Agent': head_user_agent[random.randrange(0, len(head_user_agent))],
    }  # 获得随机请求头
    return header


def url2bs(url):  # 获取网页内容
    header = requests_headers()
    try:
        f = requests.get(url, headers=header)  # 尝试获取网页
        f.raise_for_status()  # 如果状态码不是200则引发异常
        f.encoding = f.apparent_encoding  # 编码
        soup = bs4.BeautifulSoup(f.text, 'html.parser')  # 将读取到的网页解析
        return soup
    except:
        print('产生异常')  # 状态码不是200打印产生异常


def content_bv(url):
    """爬取B站月排行榜的全部100个视频的BV号，存入num_list中"""
    soup = url2bs(url)
    bvs = soup.find_all('div', class_="info")
    num_list = list()
    for bv in bvs:
        num_list.append(str(bv)[74:84])
    return num_list


proxies = ['125.66.217.114:6675', '112.251.161.82:6675',
           '117.34.253.157:6675', '113.94.72.209:6666',
           '114.105.217.144:6673', '125.92.110.80:6675',
           '112.235.126.55:6675', '14.148.99.188:6675',
           '112.240.161.20:6668', '122.82.160.148:6675',
           '175.30.224.66:6675']  # 代理IP池


# 抽取IP池IP，构建opener
def request_proxie():
    header1 = requests_headers()  # 获得随机请求头
    proxie_handler = urllib.request.ProxyHandler({'http': random.choice(proxies)})
    opener = urllib.request.build_opener(proxie_handler)
    header = []
    for key, value in header1.items():
        elem = (key, value)
        header.append(elem)
    opener.addheaders = header
    return opener


# 获取b站单个视频网页内容
def getHtmlInfo(av_num):
    opener = request_proxie()
    req_view = opener.open("https://api.bilibili.com/x/web-interface/view?bvid=" + str(av_num))
    page_view = req_view.read().decode('utf-8')
    dic_page = json.loads(page_view)  # 将获取内容转成字典形式


    if dic_page['code'] != 0:  # 判断视频是否存在，不存在返回av号及错误原因
        video_data = [av_num, dic_page['message'], dic_page['code'], '', '', '', '', '', '', '']
    else:  # 依次获取视频av号、视频标题、视频类型、上传时间、作者（UP主）、弹幕数、评论数、收藏数、投币数、点赞数
        video_info = dic_page['data']
        video_data = [video_info['bvid'],   # bv号
                      video_info['aid'],  # av号
                      video_info['title'],  # 标题
                      video_info['tname'],  # 视频类型
                      time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(video_info['ctime'])),
                      # 上传时间，网页获取时间为Unix时间戳，转换成一般时间显示模式
                      video_info['owner']['name'],  # 作者
                      video_info['stat']['danmaku'],  # 弹幕数
                      video_info['stat']['reply'],  # 评论数
                      video_info['stat']['favorite'],  # 收藏数
                      video_info['stat']['coin'],  # 投币数
                      video_info['stat']['like'],  # 点赞数
                      video_info['desc'],           # 视频简介
                      ]
        time.sleep(2 + random.randint(-1, 2))
    print("爬取bv" + str(av_num) + '已完成')  # 提示爬取进度
    req_view.close()
    return video_data


def getAllInfo(num_list):
    allinfo = list()
    for av_num in num_list:
        allinfo.append(getHtmlInfo(av_num))
    return allinfo


def saveToExcel(num_list):
    allinfo = getAllInfo(num_list)
    wb = openpyxl.Workbook()
    sheet1 = wb.create_sheet('bilibili视频信息爬取',0)
    title_ls = ['bv号', 'av号', '视频标题', '视频类型', '上传时间', '作者（UP主）', '弹幕数', '评论数', '收藏数', '投币数', '点赞数','简介']

    for i in range(len(title_ls)):  # 循环写入标题
        sheet1.cell(1, 1 + i).value = title_ls[i]
    for row in range(len(allinfo)):  # 嵌套循环写入视频信息
        print('已写入' + str(allinfo[0][0] + str(row)))  # 提示写入进度
        for column in range(len(allinfo[row])):
            sheet1.cell(row + 2, column + 1).value = allinfo[row][column]

    wb.save('./bilibili_data' + '.xlsx')
    wb.close()


if __name__ == "__main__":
    URL = 'https://www.bilibili.com/ranking/all/0/0/30'
    num_list = content_bv(URL)
    saveToExcel(num_list)
